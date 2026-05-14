#!/usr/bin/env python3
"""
Generate Excel file showing missing translations
Compares English keys with Arabic, Kurdish Sorani, and Kurdish Bahdini translations
"""

import json
import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl not installed. Installing...")
    os.system("pip3 install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

def load_translations(file_path):
    """Load translation JSON file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"⚠️  File not found: {file_path}")
        return {}
    except json.JSONDecodeError:
        print(f"❌ Invalid JSON in: {file_path}")
        return {}

def generate_excel(ar_file, ku_sorani_file, ku_bahdini_file, output_file):
    """Generate Excel file with translation comparison"""
    
    # Load all translation files
    print("📖 Loading translation files...")
    ar_trans = load_translations(ar_file)
    ku_sorani_trans = load_translations(ku_sorani_file)
    ku_bahdini_trans = load_translations(ku_bahdini_file)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    completed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Set headers
    headers = ["English (Key)", "Arabic (العربية)", "Kurdish Sorani (سۆرانی)", "Kurdish Bahdini (بادینی)", "Status"]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Track statistics
    total = len(ar_trans)
    ar_complete = 0
    sorani_complete = 0
    bahdini_complete = 0
    
    # Add data rows
    row_num = 2
    for english_key, arabic_value in sorted(ar_trans.items()):
        # Get translations
        sorani_value = ku_sorani_trans.get(english_key, "")
        bahdini_value = ku_bahdini_trans.get(english_key, "")
        
        # Determine if translations exist and are different from Arabic (not just copied)
        has_arabic = bool(arabic_value and arabic_value != english_key)
        has_sorani = bool(sorani_value and sorani_value != arabic_value)
        has_bahdini = bool(bahdini_value and bahdini_value != arabic_value)
        
        # Calculate status
        if has_sorani and has_bahdini:
            status = "✅ Complete"
        elif has_sorani or has_bahdini:
            status = "⚠️ Partial"
        else:
            status = "❌ Missing"
        
        # Update statistics
        if has_arabic:
            ar_complete += 1
        if has_sorani:
            sorani_complete += 1
        if has_bahdini:
            bahdini_complete += 1
        
        # Add row
        ws.append([
            english_key,
            arabic_value if has_arabic else "",
            sorani_value if has_sorani else "",
            bahdini_value if has_bahdini else "",
            status
        ])
        
        # Style cells based on translation status
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Highlight missing translations
            if col_num == 3 and not has_sorani:  # Kurdish Sorani
                cell.fill = missing_fill
            elif col_num == 4 and not has_bahdini:  # Kurdish Bahdini
                cell.fill = missing_fill
            elif col_num == 5 and status == "✅ Complete":
                cell.fill = completed_fill
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40  # English
    ws.column_dimensions['B'].width = 35  # Arabic
    ws.column_dimensions['C'].width = 35  # Kurdish Sorani
    ws.column_dimensions['D'].width = 35  # Kurdish Bahdini
    ws.column_dimensions['E'].width = 15  # Status
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Translation Progress Summary"])
    summary_ws.append([])
    summary_ws.append(["Language", "Translated", "Missing", "Progress %"])
    summary_ws.append([
        "Arabic (العربية)",
        ar_complete,
        total - ar_complete,
        f"{(ar_complete/total*100):.1f}%"
    ])
    summary_ws.append([
        "Kurdish Sorani (سۆرانی)",
        sorani_complete,
        total - sorani_complete,
        f"{(sorani_complete/total*100):.1f}%"
    ])
    summary_ws.append([
        "Kurdish Bahdini (بادینی)",
        bahdini_complete,
        total - bahdini_complete,
        f"{(bahdini_complete/total*100):.1f}%"
    ])
    
    # Style summary
    summary_ws['A1'].font = Font(bold=True, size=14)
    for row in range(3, 7):
        for col in range(1, 5):
            cell = summary_ws.cell(row=row, column=col)
            if row == 3:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 15
    summary_ws.column_dimensions['D'].width = 15
    
    # Save workbook
    wb.save(output_file)
    
    print(f"\n✅ Excel file generated: {output_file}")
    print(f"\n📊 Translation Statistics:")
    print(f"   Total terms: {total}")
    print(f"   Arabic: {ar_complete}/{total} ({ar_complete/total*100:.1f}%)")
    print(f"   Kurdish Sorani: {sorani_complete}/{total} ({sorani_complete/total*100:.1f}%)")
    print(f"   Kurdish Bahdini: {bahdini_complete}/{total} ({bahdini_complete/total*100:.1f}%)")
    print(f"   Missing Sorani: {total - sorani_complete}")
    print(f"   Missing Bahdini: {total - bahdini_complete}")

if __name__ == "__main__":
    ar_file = "resources/lang/ar.json"
    ku_sorani_file = "resources/lang/ku-sorani.json"
    ku_bahdini_file = "resources/lang/ku-bahdini.json"
    output_file = "ConCure_Translations.xlsx"
    
    generate_excel(ar_file, ku_sorani_file, ku_bahdini_file, output_file)
